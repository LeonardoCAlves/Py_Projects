import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from datetime import date
import pandas as pd


# Criando base de dados
def criando_arquivo_base():
    colunas = ["Nome", "dt_nascimento", "CPF", "Telefone", "Email", "Endereço",
               "Profissão", "Sexo", "Produto", "dt_cadastro"]
    
    df = pd.DataFrame(columns=colunas)
    df.to_excel('dados.xlsx', index=False, sheet_name='Base_Clientes')

# Lendo base de dados
def titulos():
    workbook = pd.read_excel('dados.xlsx', sheet_name='Base_Clientes')
    titles = list(
        workbook.columns
    )
    return titles

# Inserindo registros
def inserindo_registros(cliente):

    # carregando pasta de trabalho
    workbook = openpyxl.load_workbook('dados.xlsx')

    # selecionando a planilha Base_Clientes
    worksheet = workbook['Base_Clientes']

    # encontrando a última linha preenchida
    ultima_linha = worksheet.max_row

    while all([
            worksheet.cell(row=ultima_linha, column=column_index_from_string('A')).value is None,
            worksheet.cell(row=ultima_linha, column=column_index_from_string('B')).value is None,
            worksheet.cell(row=ultima_linha, column=column_index_from_string('C')).value is None,
            worksheet.cell(row=ultima_linha, column=column_index_from_string('D')).value is None,
            worksheet.cell(row=ultima_linha, column=column_index_from_string('E')).value is None,
            worksheet.cell(row=ultima_linha, column=column_index_from_string('F')).value is None,
            worksheet.cell(row=ultima_linha, column=column_index_from_string('G')).value is None,
            worksheet.cell(row=ultima_linha, column=column_index_from_string('H')).value is None,
            worksheet.cell(row=ultima_linha, column=column_index_from_string('I')).value is None,
            worksheet.cell(row=ultima_linha, column=column_index_from_string('J')).value is None
        ]):

        ultima_linha -= 1

    # incrementando o número da linha em 1 para começar a inserir na próxima vazia
    linha = ultima_linha + 1

    # recebendo dados do cliente
    dados_cliente = cliente

    # inserindo dados na planilha 
    worksheet.cell(row=linha, column=column_index_from_string('A'), value = dados_cliente[0])
    worksheet.cell(row=linha, column=column_index_from_string('B'), value = dados_cliente[1])
    worksheet.cell(row=linha, column=column_index_from_string('C'), value = dados_cliente[2])
    worksheet.cell(row=linha, column=column_index_from_string('D'), value = dados_cliente[3])
    worksheet.cell(row=linha, column=column_index_from_string('E'), value = dados_cliente[4])
    worksheet.cell(row=linha, column=column_index_from_string('F'), value = dados_cliente[5])
    worksheet.cell(row=linha, column=column_index_from_string('G'), value = dados_cliente[6])
    worksheet.cell(row=linha, column=column_index_from_string('H'), value = dados_cliente[7])
    worksheet.cell(row=linha, column=column_index_from_string('I'), value = dados_cliente[8])
    worksheet.cell(row=linha, column=column_index_from_string('J'), value = dados_cliente[9])

    # salvando a pasta de trabalho atualizada
    workbook.save('dados.xlsx')


registro = ["Léo", "01/11/1993", "12312312312", "11979797979", "l@email.com", "Ruas das Coves, 30, SP",
               "Programador", "Masculino", "Camisa do Mengão", f"{date.today().strftime('%d/%m/%Y')}"]

inserindo_registros(registro)